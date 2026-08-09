"""
utils/excel_export.py
======================
Genera el parte de trabajo mensual en formato .xlsx, como alternativa
al PDF cuando el destino son otras hojas de calculo (por ejemplo, si
administracion quiere seguir consolidando los datos en Excel).

A diferencia de la plantilla original, aqui el Excel se genera siempre
a partir de los datos guardados en SQLite -- no hay formulas manuales
que se puedan desajustar de una hoja a otra.
"""

from __future__ import annotations

import io
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database.db import MESES_ES
from utils.calculos import resumen_de_partes

COLOR_CABECERA = "1F3B57"       # azul corporativo (mismo tono que el PDF)
COLOR_CABECERA_TEXTO = "FFFFFF"
COLOR_RESUMEN = "DCE6EF"

COLUMNAS = [
    ("Fecha", 12), ("Tipo", 12), ("Horas", 9), ("H. Extra", 10),
    ("Dietas", 9), ("Cliente", 24), ("Dirección", 28), ("Nº Trabajo", 14),
    ("Trabajo realizado", 45), ("Observaciones / Problemas", 45), ("Compañero", 18),
]


def generar_excel_mensual(
    anio: int,
    mes: int,
    nombre_trabajador: str,
    partes: list[sqlite3.Row],
) -> bytes:
    """
    Construye el libro Excel del parte mensual y devuelve sus bytes
    (listos para `st.download_button`), con una hoja: tabla de dias +
    fila de resumen, igual que la version en PDF.
    """
    wb = Workbook()
    ws = wb.active
    nombre_mes = MESES_ES[mes - 1]
    ws.title = nombre_mes[:31]  # Excel limita el nombre de hoja a 31 caracteres

    # --- Titulo -----------------------------------------------------
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Parte de Trabajo Mensual - {nombre_mes} {anio} - {nombre_trabajador or 'Trabajador/a'}"
    ws["A1"].font = Font(size=14, bold=True)

    # --- Cabecera de la tabla ----------------------------------------
    fila_cabecera = 3
    for col_idx, (titulo, ancho) in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=fila_cabecera, column=col_idx, value=titulo)
        celda.font = Font(bold=True, color=COLOR_CABECERA_TEXTO)
        celda.fill = PatternFill("solid", fgColor=COLOR_CABECERA)
        celda.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # --- Filas de datos ------------------------------------------------
    fila = fila_cabecera + 1
    for p in partes:
        valores = [
            p["fecha"], p["tipo_jornada"],
            p["horas_jornada"] or None, p["horas_extra"] or None,
            p["dietas"] or None, p["cliente"] or "", p["direccion"] or "",
            p["numero_trabajo"] or "",
            p["trabajo_realizado"] or "", p["observaciones"] or "",
            p["companero"] or "",
        ]
        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=col_idx, value=valor)
            celda.alignment = Alignment(vertical="top", wrap_text=True)
        fila += 1

    # --- Resumen del mes -------------------------------------------------
    resumen = resumen_de_partes(partes)
    fila += 1
    etiquetas = ["Horas trabajadas", "Horas extra", "Dietas",
                 "Días vacaciones", "Días de baja", "Días de guardia"]
    valores_resumen = [
        resumen["horas_totales"], resumen["horas_extra"], resumen["dietas"],
        resumen["dias_vacaciones"], resumen["dias_baja"], resumen["dias_guardia"],
    ]
    for col_idx, (etiqueta, valor) in enumerate(zip(etiquetas, valores_resumen), start=1):
        celda_etiqueta = ws.cell(row=fila, column=col_idx, value=etiqueta)
        celda_etiqueta.font = Font(bold=True)
        celda_etiqueta.fill = PatternFill("solid", fgColor=COLOR_RESUMEN)
        celda_valor = ws.cell(row=fila + 1, column=col_idx, value=valor)
        celda_valor.fill = PatternFill("solid", fgColor=COLOR_RESUMEN)
        celda_valor.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(row=fila_cabecera + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
